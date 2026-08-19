import io
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.deps import (
    check_customer_access,
    check_customer_write_access,
    get_current_active_user,
    get_db,
)
from app.models.customer import Customer
from app.models.user import User
from app.schemas.customer import (
    BatchAssign,
    BatchStatus,
    CustomerCreate,
    CustomerResponse,
    CustomerUpdate,
    StageAdvance,
    StatusUpdate,
)
from app.services.customer_service import (
    batch_assign,
    batch_status,
    create_customer,
    get_customer_detail,
    get_timeline,
    search_customers,
    update_customer,
)
from app.services.stage_service import advance_stage, get_alert_level, get_stay_days

router = APIRouter(prefix="/api/v1/customers", tags=["Customer Management"])


def _build_response(customer: Customer) -> dict:
    """Build unified customer response"""
    data = CustomerResponse.model_validate(customer).model_dump()
    data["stay_days"] = get_stay_days(customer)
    data["alert_level"] = get_alert_level(customer)
    data["sales_name"] = customer.sales.name if customer.sales else None
    return data


@router.post("", response_model=dict)
def create(
    data: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Create customer"""
    customer = create_customer(db, data.model_dump(), current_user.id)
    return {
        "code": 0,
        "message": "success",
        "data": _build_response(customer),
    }


@router.get("", response_model=dict)
def list_customers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = Query(None),
    stage: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    sales_id: Optional[UUID] = Query(None),
    region: Optional[str] = Query(None),
    alert_level: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Customer list (pagination + filters)"""
    # sales/cs can only see their own customers; admin/pm see everything
    if current_user.role in ("sales", "cs"):
        sales_id = current_user.id
    customers, total = search_customers(
        db, keyword or "", page, page_size, stage, status, sales_id, region, alert_level
    )
    items = [_build_response(c) for c in customers]
    return {
        "code": 0,
        "message": "success",
        "data": {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/advanced-search", response_model=dict)
def advanced_search(
    name: str | None = Query(None),
    contact_person: str | None = Query(None),
    phone: str | None = Query(None),
    company: str | None = Query(None),
    region: str | None = Query(None),
    source_channel: str | None = Query(None),
    current_stage: int | None = Query(None, ge=1, le=8),
    status: str | None = Query(None),
    sales_id: Optional[UUID] = Query(None),
    min_contract_amount: float | None = Query(None, ge=0),
    max_contract_amount: float | None = Query(None, ge=0),
    created_after: str | None = Query(None),
    created_before: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Advanced search with multi-field filters"""
    from datetime import datetime
    from sqlalchemy import and_

    query = db.query(Customer).filter(Customer.status != "deleted")

    # sales/cs can only search their own customers; admin/pm see everything
    if current_user.role in ("sales", "cs"):
        query = query.filter(Customer.sales_id == current_user.id)

    if name:
        query = query.filter(Customer.name.ilike(f"%{name}%"))
    if contact_person:
        query = query.filter(Customer.contact_person.ilike(f"%{contact_person}%"))
    if phone:
        query = query.filter(Customer.phone.ilike(f"%{phone}%"))
    if company:
        query = query.filter(Customer.company.ilike(f"%{company}%"))
    if region:
        query = query.filter(Customer.region == region)
    if source_channel:
        query = query.filter(Customer.source_channel == source_channel)
    if current_stage:
        query = query.filter(Customer.current_stage == current_stage)
    if status:
        query = query.filter(Customer.status == status)
    if sales_id:
        query = query.filter(Customer.sales_id == sales_id)
    if min_contract_amount is not None:
        query = query.filter(Customer.contract_amount >= min_contract_amount)
    if max_contract_amount is not None:
        query = query.filter(Customer.contract_amount <= max_contract_amount)
    if created_after:
        query = query.filter(Customer.created_at >= datetime.fromisoformat(created_after))
    if created_before:
        query = query.filter(Customer.created_at <= datetime.fromisoformat(created_before))

    total = query.count()
    customers = (
        query.order_by(Customer.updated_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    from app.services.stage_service import get_alert_level, get_stay_days

    items = []
    for c in customers:
        items.append({
            "id": str(c.id),
            "name": c.name,
            "contact_person": c.contact_person,
            "phone": c.phone,
            "company": c.company,
            "region": c.region,
            "source_channel": c.source_channel,
            "current_stage": c.current_stage,
            "status": c.status,
            "contract_amount": float(c.contract_amount or 0),
            "sales_id": str(c.sales_id) if c.sales_id else None,
            "stay_days": get_stay_days(c),
            "alert_level": get_alert_level(c),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        })

    return {
        "code": 0,
        "message": "success",
        "data": {"items": items, "total": total, "page": page, "page_size": page_size},
    }


@router.get("/export")
def export_customers(
    stage: Optional[int] = Query(None),
    sales_id: Optional[UUID] = Query(None),
    region: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export customers to Excel file"""
    from datetime import datetime

    query = db.query(Customer).filter(Customer.status != "deleted")
    # sales/cs can only export their own customers; admin/pm export everything
    if current_user.role in ("sales", "cs"):
        query = query.filter(Customer.sales_id == current_user.id)
    if stage is not None:
        query = query.filter(Customer.current_stage == stage)
    if sales_id is not None:
        query = query.filter(Customer.sales_id == sales_id)
    if region is not None:
        query = query.filter(Customer.region == region)
    if start_date:
        query = query.filter(Customer.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Customer.created_at <= datetime.fromisoformat(end_date))

    customers = query.order_by(Customer.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "name", "contact_person", "phone", "wechat", "email",
        "company", "region", "source_channel", "current_stage",
        "contract_amount", "created_at",
    ]
    ws.append(headers)

    for c in customers:
        ws.append([
            c.name,
            c.contact_person,
            c.phone,
            c.wechat,
            c.email,
            c.company,
            c.region,
            c.source_channel,
            c.current_stage,
            float(c.contract_amount) if c.contract_amount else 0,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=customers.xlsx",
        },
    )


@router.get("/export-template")
def export_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Download Excel import template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "name", "contact_person", "phone", "wechat", "email",
        "company", "region", "source_channel",
    ]
    ws.append(headers)
    ws.append(["Example Customer", "John Doe", "13800000000", "wechat001", "demo@example.com", "Example Corp", "East China", "Online Ads"])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=customer_import_template.xlsx",
        },
    )


@router.get("/{id}", response_model=dict)
def get_detail(
    id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Customer detail"""
    customer = get_customer_detail(db, id)
    check_customer_access(customer, current_user)
    return {
        "code": 0,
        "message": "success",
        "data": _build_response(customer),
    }


@router.put("/{id}", response_model=dict)
def update(
    id: UUID,
    data: CustomerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Update customer"""
    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    check_customer_write_access(customer, current_user)
    customer = update_customer(db, id, data.model_dump(exclude_none=True), current_user.id)
    return {
        "code": 0,
        "message": "success",
        "data": _build_response(customer),
    }


@router.delete("/{id}", response_model=dict)
def delete(
    id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Soft delete customer"""
    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Permission denied: only admins can delete customers")
    customer.status = "deleted"
    db.commit()
    from app.services.audit_service import log_action
    log_action(
        db,
        current_user.id,
        "delete",
        "customer",
        object_id=customer.id,
        customer_id=customer.id,
        before_data={"name": customer.name, "status": customer.status},
    )
    return {"code": 0, "message": "success", "data": None}


@router.put("/{id}/stage", response_model=dict)
def advance(
    id: UUID,
    data: StageAdvance,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Advance customer stage"""
    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    check_customer_write_access(customer, current_user)
    from_stage = customer.current_stage
    customer = advance_stage(db, id, data.new_stage, current_user.id, data.remark)
    from app.services.audit_service import log_action
    log_action(
        db,
        current_user.id,
        "advance_stage",
        "customer",
        object_id=customer.id,
        customer_id=customer.id,
        before_data={"current_stage": from_stage},
        after_data={"current_stage": customer.current_stage, "remark": data.remark},
    )
    return {
        "code": 0,
        "message": "success",
        "data": _build_response(customer),
    }


@router.put("/{id}/status", response_model=dict)
def update_status(
    id: UUID,
    data: StatusUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Update customer status"""
    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    check_customer_write_access(customer, current_user)
    from app.services.audit_service import log_action
    old_status = customer.status
    customer.status = data.status
    customer.lost_reason = data.lost_reason
    db.commit()
    db.refresh(customer)
    log_action(
        db,
        current_user.id,
        "update_status",
        "customer",
        object_id=customer.id,
        customer_id=customer.id,
        before_data={"status": old_status},
        after_data={"status": customer.status, "lost_reason": customer.lost_reason},
    )
    return {
        "code": 0,
        "message": "success",
        "data": _build_response(customer),
    }


@router.put("/{id}/assign", response_model=dict)
def assign(
    id: UUID,
    data: BatchAssign,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Transfer sales person (single)"""
    _assert_batch_customer_access(db, [id], current_user)
    count = batch_assign(db, [id], data.new_sales_id)
    return {"code": 0, "message": "success", "data": {"affected": count}}


@router.put("/{id}/rollback", response_model=dict)
def rollback(
    id: UUID,
    data: StageAdvance,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Rollback customer to previous stage (admin only)"""
    from app.core.deps import check_role_admin
    check_role_admin(current_user)
    from app.services.stage_service import rollback_stage
    customer = rollback_stage(db, id, current_user.id, data.remark)
    return {
        "code": 0,
        "message": "success",
        "data": {
            "id": str(customer.id),
            "name": customer.name,
            "current_stage": customer.current_stage,
            "stage_entered_at": customer.stage_entered_at.isoformat(),
            "status": customer.status,
        },
    }


@router.get("/{id}/timeline", response_model=dict)
def timeline(
    id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Customer timeline"""
    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    check_customer_access(customer, current_user)
    items = get_timeline(db, id)
    return {"code": 0, "message": "success", "data": items}


def _assert_batch_customer_access(db, customer_ids: list, current_user) -> None:
    """Admins may operate on any customer; other roles only on their own customers."""
    if current_user.role == "admin":
        return
    customers = db.query(Customer).filter(Customer.id.in_(customer_ids)).all()
    for c in customers:
        if c.sales_id != current_user.id:
            raise HTTPException(status_code=403, detail="Permission denied")


@router.post("/batch/assign", response_model=dict)
def batch_assign_route(
    data: BatchAssign,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Batch transfer sales"""
    _assert_batch_customer_access(db, data.customer_ids, current_user)
    count = batch_assign(db, data.customer_ids, data.new_sales_id)
    return {"code": 0, "message": "success", "data": {"affected": count}}


@router.post("/batch/status", response_model=dict)
def batch_status_route(
    data: BatchStatus,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Batch update status"""
    _assert_batch_customer_access(db, data.customer_ids, current_user)
    count = batch_status(db, data.customer_ids, data.status)
    return {"code": 0, "message": "success", "data": {"affected": count}}


@router.post("/batch/delete", response_model=dict)
def batch_delete(
    ids: list[UUID],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Batch delete (soft delete)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Permission denied: only admins can delete customers")
    _assert_batch_customer_access(db, ids, current_user)
    count = (
        db.query(Customer)
        .filter(Customer.id.in_(ids))
        .update({"status": "deleted"}, synchronize_session=False)
    )
    db.commit()
    return {"code": 0, "message": "success", "data": {"affected": count}}


@router.get("/{id}/export")
def export_single_customer(
    id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export single customer details as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    customer = db.query(Customer).filter(Customer.id == id).first()
    if not customer:
        return {"code": 4004, "message": "Customer not found"}

    # Object-level access enforcement (admin: full; sales: own customers only)
    check_customer_access(customer, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Detail"

    # Header
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Basic info
    ws.cell(row=1, column=1, value="Customer Detail").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    info_rows = [
        ("Name", customer.name or ""),
        ("Contact Person", customer.contact_person or ""),
        ("Phone", customer.phone or ""),
        ("Wechat", customer.wechat or ""),
        ("Email", customer.email or ""),
        ("Company", customer.company or ""),
        ("Region", customer.region or ""),
        ("Source Channel", customer.source_channel or ""),
        ("Current Stage", str(customer.current_stage)),
        ("Status", customer.status or ""),
        ("Contract Amount", float(customer.contract_amount or 0)),
        ("Paid Amount", float(customer.paid_amount or 0)),
    ]

    for i, (label, value) in enumerate(info_rows, start=3):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_a.font = header_font
        cell_a.border = thin_border
        cell_b = ws.cell(row=i, column=2, value=value)
        cell_b.border = thin_border
        ws.merge_cells(f"B{i}:D{i}")

    # Task info if exists
    from app.models.task import Task
    tasks = db.query(Task).filter(Task.customer_id == id).all()
    if tasks:
        row = len(info_rows) + 5
        ws.cell(row=row, column=1, value="Tasks").font = Font(bold=True, size=12)
        row += 1
        for t in tasks:
            ws.cell(row=row, column=1, value=t.name).border = thin_border
            ws.cell(row=row, column=2, value=t.status).border = thin_border
            ws.cell(row=row, column=3, value=str(t.due_date or "")).border = thin_border
            row += 1

    # Payment info if exists
    from app.models.payment import Payment
    payments = db.query(Payment).filter(Payment.customer_id == id).all()
    if payments:
        row += 1
        ws.cell(row=row, column=1, value="Payments").font = Font(bold=True, size=12)
        row += 1
        for p in payments:
            ws.cell(row=row, column=1, value=str(p.amount)).border = thin_border
            ws.cell(row=row, column=2, value=str(p.payment_date or "")).border = thin_border
            ws.cell(row=row, column=3, value=p.payment_type).border = thin_border
            row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"customer_{customer.name}_{customer.id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import", response_model=dict)
async def import_customers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Excel import - parse Excel file and create customer records"""
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip() if row[0] else None
        contact_person = str(row[1]).strip() if len(row) > 1 and row[1] else None
        phone = str(row[2]).strip() if len(row) > 2 and row[2] else None

        if not name or not phone:
            continue

        # Skip rows whose phone already exists (avoid MySQL unique-constraint 500)
        existing = (
            db.query(Customer)
            .filter(Customer.phone == phone, Customer.status != "deleted")
            .first()
        )
        if existing:
            skipped += 1
            continue

        customer = Customer(
            name=name,
            contact_person=contact_person,
            phone=phone,
            wechat=str(row[3]).strip() if len(row) > 3 and row[3] else None,
            email=str(row[4]).strip() if len(row) > 4 and row[4] else None,
            company=str(row[5]).strip() if len(row) > 5 and row[5] else None,
            region=str(row[6]).strip() if len(row) > 6 and row[6] else None,
            source_channel=str(row[7]).strip() if len(row) > 7 and row[7] else None,
            sales_id=current_user.id,
        )
        db.add(customer)
        count += 1

    db.commit()
    return {"code": 0, "message": "success", "data": {"imported": count, "skipped": skipped}}
